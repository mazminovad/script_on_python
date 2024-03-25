#Монтируем библиотеки
import psycopg2 
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


#Подключение к БД
conn = psycopg2.connect(
    host="localhost",
    port=****,
    database="p_prod",
    user="******",
    password="*******"
)

#Запросы к БД 
cur1 = conn.cursor()
cur1.execute("""select a."Id" as "Account id",
a."ExternalAccountId",
a."Phone",
o."Id"as "Order id",
o."UniqId" as "Uniq order id",
o."CreatedAt",
o."UpdatedAt",
cast(o."ExpirationDate" AT TIME ZONE 'UTC' AS timestamp),
cast(o."CancellationDate" AT TIME ZONE 'UTC' AS timestamp),
o."Status",
o."CancellationByBuyerId",
o."CancellationByPharmacyId",
o."CancellationBySystemId",
oc."Code",
oc."RootCause",
o."PharmacyName",
ppl."FullAddress",
pp."ExternalId" as "External pharmacy id",
pp."SystemPublicId" as "System Public pharmacy Id",
pp."ExternalId" as "External pharmacy Id",
pp."ExternalPublicId" as "External public pharmacy Id",
pp."ExternalPartnerId",
pp."ExternalPharmacyNetworkId"
from orders o
join accounts a on a."Id" = o."AccountId"
join partners_pharmacies pp on o."PharmacyId" = pp."Id"
join partners_pharmacies_location ppl on pp."Id" = ppl."PharmacyId"
left join orders_cancellation oc on oc."Id" = o."CancellationByPharmacyId" 
or oc."Id" = o."CancellationByBuyerId" or oc."Id" = o."CancellationBySystemId"
WHERE o."CreatedAt" >= date_trunc('month', now() - interval '1 month')  -- начало предыдущего месяца
AND o."CreatedAt" < date_trunc('month', now())  -- начало текущего месяца
""")
rows1 = cur1.fetchall()


#Создаем файл exls и добавляем листы
wb = Workbook()
ws1 = wb.active
ws1.title = "report"


#Добавление orders
ws1['A1'] = 'Id'             
ws1['B1'] = 'ExternalAccountId'
ws1['C1'] = 'Phone'
ws1['D1'] = 'Order id'
ws1['E1'] = 'Uniq order id'
ws1['F1'] = 'CreatedAt'
ws1['G1'] = 'UpdatedAt'
ws1['H1'] = 'ExpirationDate'
ws1['I1'] = 'Order cancellation date'
ws1['J1'] = 'Status'
ws1['K1'] = 'CancellationByBuyerId'
ws1['L1'] = 'CancellationByPharmacyId'
ws1['M1'] = 'CancellationBySystemId'
ws1['N1'] = 'Code'
ws1['O1'] = 'RootCause'
ws1['P1'] = 'PharmacyName'
ws1['Q1'] = 'FullAddress'
ws1['R1'] = 'External pharmacy id'
ws1['S1'] = 'System Public pharmacy Id'
ws1['T1'] = 'External pharmacy Id'
ws1['U1'] = 'External public pharmacy Id'
ws1['V1'] = 'ExternalPartnerId'
ws1['W1'] = 'ExternalPharmacyNetworkId'
#ws1['X1'] = 'Public good Id'
#ws1['Y1'] = 'GoodId'
#ws1['Z1'] = 'GoodName'
#ws1['AA1'] = 'PharmacyGoodId'
#ws1['AB1'] = 'PharmacyGoodName'
#ws1['AC1'] = 'Quantity'
#ws1['AD1'] = 'Price'



#Добавление данных из orders
row_num = 2
for row in rows1:
    ws1.cell(row=row_num, column=1).value = row[0]
    ws1.cell(row=row_num, column=2).value = row[1]
    ws1.cell(row=row_num, column=3).value = row[2]
    ws1.cell(row=row_num, column=4).value = row[3]
    ws1.cell(row=row_num, column=5).value = row[4]
    ws1.cell(row=row_num, column=6).value = row[5]
    ws1.cell(row=row_num, column=7).value = row[6]
    ws1.cell(row=row_num, column=8).value = row[7]
    ws1.cell(row=row_num, column=9).value = row[8]
    ws1.cell(row=row_num, column=10).value = row[9]
    ws1.cell(row=row_num, column=11).value = row[10]
    ws1.cell(row=row_num, column=12).value = row[11]
    ws1.cell(row=row_num, column=13).value = row[12]
    ws1.cell(row=row_num, column=14).value = row[13]
    ws1.cell(row=row_num, column=15).value = row[14]
    ws1.cell(row=row_num, column=16).value = row[15]
    ws1.cell(row=row_num, column=17).value = row[16]
    ws1.cell(row=row_num, column=18).value = row[17]
    ws1.cell(row=row_num, column=19).value = row[18]
    ws1.cell(row=row_num, column=20).value = row[19]
    ws1.cell(row=row_num, column=21).value = row[20]
    ws1.cell(row=row_num, column=22).value = row[21]
    ws1.cell(row=row_num, column=23).value = row[22]
    #ws1.cell(row=row_num, column=24).value = row[23]
    #ws1.cell(row=row_num, column=25).value = row[24]
    #ws1.cell(row=row_num, column=26).value = row[25]
    #ws1.cell(row=row_num, column=27).value = row[26]
    #ws1.cell(row=row_num, column=28).value = row[27]
    #ws1.cell(row=row_num, column=29).value = row[28]
    #ws1.cell(row=row_num, column=30).value = row[29]
    row_num += 1



#Сохранение файла
wb.save('report_analyst.xlsx')

# Отправка файла на почту
msg = MIMEMultipart()
msg['From'] = 'xxxxxx@e.ru'
msg['To'] = 'xxxxxx@m.ru'
msg['Subject'] = 'Отчет для аналитики'

with open("report_analyst.xlsx", "rb") as attachment:
    xls_part = MIMEApplication(
        attachment.read(),
        Name="report_analyst.xlsx"
    )
    xls_part['Content-Disposition'] = 'attachment; filename="report_analyst.xlsx"'
    msg.attach(xls_part)

server = smtplib.SMTP('smtp.mail.ru', 25)
server.starttls()
server.login('xxxxxx@e.ru', '**********')
text = msg.as_string()
server.sendmail(msg['From'],msg['To'], text)
server.quit()




